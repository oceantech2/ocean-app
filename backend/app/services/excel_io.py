"""Importação/exportação de NFs, Colaboradores e Contas a Pagar via xlsx.

Funções puras (sem FastAPI/DB) que recebem bytes e devolvem listas de dicts prontos
para validação/persistência nas rotas. Os templates NUNCA são sobrescritos em disco —
o preenchimento é feito sempre em memória (BytesIO) e devolvido como bytes.

Templates usados:
  - modelo_nfs_colaboradores.xlsx: aba "Entradas" para NFs
  - modelo_colaboradores.xlsx: aba "Colaboradores" simplificada (sem férias)
  - modelo_contas.xlsx: aba "JAN26" para Contas a Pagar / Fluxo de Caixa
"""
import io
import os
from datetime import date, datetime
from typing import Any

import openpyxl

_TEMPLATES = os.path.join(os.path.dirname(__file__), "..", "templates")
TEMPLATE_NFS_PATH   = os.path.join(_TEMPLATES, "modelo_nfs_colaboradores.xlsx")
TEMPLATE_COLAB_PATH = os.path.join(_TEMPLATES, "modelo_colaboradores.xlsx")
TEMPLATE_CONTAS_PATH = os.path.join(_TEMPLATES, "modelo_contas.xlsx")

# ==================== NFs ====================
NF_SHEET = "Entradas"
# Linha do cabeçalho no template → dados a partir de NF_DATA_START_ROW
NF_HEADER_ROW = 2
NF_DATA_START_ROW = 3
# Índices de coluna no template (1-based)
NF_COLS = {
    "numero":         2,
    "razao_social":   3,
    "posicao":        4,
    "valor_bruto":    5,
    "valor_liquido":  6,
    "data_emissao":   8,
    "data_vencimento": 9,
    "data_pagamento": 10,
    "lead":           11,
    "conducao":       12,
    "placement":      13,
}

# Aliases aceitos no cabeçalho ao importar (lowercase)
NF_COL_ALIASES = [
    ("numero",          ["nf", "número", "numero", "nº"]),
    ("razao_social",    ["empresa", "razão social", "razao social", "cliente"]),
    ("posicao",         ["posição", "posicao", "cargo", "vaga"]),
    ("valor_bruto",     ["valor", "valor bruto", "valor total"]),
    ("valor_liquido",   ["valor líquido", "valor liquido", "vlr líquido", "vlr liquido"]),
    ("data_emissao",    ["data de emissão", "data de emissao", "data emissão", "data emissao", "data de emsisão"]),
    ("data_vencimento", ["data vencimento", "data de vencimento", "vencimento"]),
    ("data_pagamento",  ["data pagamento", "data de pagamento", "pagamento"]),
    ("lead",            ["lead"]),
    ("conducao",        ["condução", "conducao", "conducão"]),
    ("placement",       ["placement"]),
]

# ==================== Colaboradores ====================
COL_SHEET = "Colaboradores"
COL_HEADER_ROW = 1       # header na linha 1 no novo template
COL_DATA_START_ROW = 2   # dados a partir da linha 2
COLAB_COLS = {
    "nome":             2,
    "cpf":              3,
    "cargo":            4,
    "salario":          5,
    "data_nascimento":  6,
    "endereco":         7,
    "data_admissao":    8,
    "data_desligamento": 9,
}

COLAB_COL_ALIASES = [
    ("nome",              ["nome", "nome "]),
    ("cpf",               ["cpf"]),
    ("cargo",             ["posição", "posicao", "cargo"]),
    ("salario",           ["salário", "salario"]),
    ("data_nascimento",   ["nascimento", "data nascimento", "data de nascimento"]),
    ("endereco",          ["endereço", "endereco"]),
    ("data_admissao",     ["data de inicio", "data inicio", "data admissão", "data admissao"]),
    ("data_desligamento", ["data desligamento", "data de desligamento"]),
]

# Blocos de férias no template combinado (modelo_nfs_colaboradores.xlsx)
FERIAS_BLOCKS = [
    (11, 12, 13),
    (15, 16, 17),
    (19, 20, 21),
    (23, 24, 25),
    (27, 28, 29),
]
DIAS_DIREITO_DEFAULT = 30

# ==================== Contas a Pagar / Fluxo de Caixa ====================
CONTAS_DATA_START_ROW = 6   # linhas 1-5 são título/cabeçalhos/saldo anterior
CONTAS_COLS = {
    "data":       1,
    "descricao":  2,
    "forma_pgto": 3,
    "debito":     4,   # saídas (contas a pagar)
    "credito":    5,   # entradas (NFs recebidas) — ignorado no import de contas
}

MESES_NOME = {
    1: "JAN", 2: "FEV", 3: "MAR", 4: "ABR", 5: "MAI", 6: "JUN",
    7: "JUL", 8: "AGO", 9: "SET", 10: "OUT", 11: "NOV", 12: "DEZ",
}


# ==================== Helpers ====================
def _to_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        v = value.strip()
        if not v:
            return None
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(v, fmt).date()
            except ValueError:
                continue
    return None


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        v = value.strip().replace("R$", "").replace(".", "").replace(",", ".").strip()
        if not v:
            return None
        try:
            return float(v)
        except ValueError:
            return None
    return None


def _to_int(value: Any) -> int | None:
    f = _to_float(value)
    return int(f) if f is not None else None


def _to_str(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


# ==================== PARSING: NFs ====================
def _detect_nf_header(ws) -> tuple[int, dict]:
    """Detecta linha de cabeçalho (varre linhas 1-5) pela coluna 'NF'."""
    for header_row in range(1, 6):
        row_vals = {
            c: (str(ws.cell(row=header_row, column=c).value or "")).strip().lower()
            for c in range(1, ws.max_column + 1)
        }
        col_map: dict[str, int] = {}
        for campo, aliases in NF_COL_ALIASES:
            for col, val in row_vals.items():
                if val in aliases:
                    col_map[campo] = col
                    break
        if "numero" in col_map:
            return header_row, col_map
    raise ValueError(
        "Coluna 'NF' não encontrada nas primeiras 5 linhas. "
        "Verifique se o arquivo tem a aba 'Entradas' com o cabeçalho correto."
    )


def parse_nfs_xlsx(file_bytes: bytes) -> list[dict]:
    """Lê a aba 'Entradas' detectando layout automaticamente.

    Linha com número preenchido e razão social vazia → NF cancelada.
    """
    wb_vals = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    if NF_SHEET not in wb_vals.sheetnames:
        raise ValueError(f"Aba '{NF_SHEET}' não encontrada no arquivo")
    ws_vals = wb_vals[NF_SHEET]

    header_row, col_map = _detect_nf_header(ws_vals)
    data_start = header_row + 1

    def cell_val(row, campo):
        col = col_map.get(campo)
        return ws_vals.cell(row=row, column=col).value if col else None

    nf_count: dict[str, int] = {}
    registros = []
    for row in range(data_start, ws_vals.max_row + 1):
        nf_raw = cell_val(row, "numero")
        if nf_raw is None or str(nf_raw).strip() == "":
            continue

        numero_base = str(int(nf_raw)) if isinstance(nf_raw, float) else str(nf_raw).strip()
        empresa = _to_str(cell_val(row, "razao_social"))

        # Número preenchido + razão social vazia → cancelada
        is_cancelada = not empresa
        if is_cancelada:
            empresa = "Cancelada"

        nf_count[numero_base] = nf_count.get(numero_base, 0) + 1
        ocorrencia = nf_count[numero_base]

        registros.append({
            "_linha": row,
            "numero": numero_base,
            "duplicado_arquivo": ocorrencia > 1,
            "razao_social": empresa,
            "cancelada": is_cancelada,
            "posicao": _to_str(cell_val(row, "posicao")),
            "valor_bruto": _to_float(cell_val(row, "valor_bruto")) or 0.0,
            "valor_liquido": _to_float(cell_val(row, "valor_liquido")) or 0.0,
            "data_emissao": _to_date(cell_val(row, "data_emissao")),
            "data_vencimento": _to_date(cell_val(row, "data_vencimento")),
            "data_pagamento": _to_date(cell_val(row, "data_pagamento")),
            "lead_nome": _to_str(cell_val(row, "lead")),
            "conducao_nome": _to_str(cell_val(row, "conducao")),
            "placement_nome": _to_str(cell_val(row, "placement")),
        })

    return registros


# ==================== PARSING: Colaboradores ====================
def _detect_colab_header(ws) -> tuple[int, dict]:
    """Detecta linha de cabeçalho da aba Colaboradores."""
    for header_row in range(1, 6):
        row_vals = {
            c: (str(ws.cell(row=header_row, column=c).value or "")).strip().lower()
            for c in range(1, ws.max_column + 1)
        }
        col_map: dict[str, int] = {}
        for campo, aliases in COLAB_COL_ALIASES:
            for col, val in row_vals.items():
                if val in aliases and campo not in col_map:
                    col_map[campo] = col
                    break
        if "nome" in col_map and "cpf" in col_map:
            return header_row, col_map
    raise ValueError("Colunas 'Nome' e 'CPF' não encontradas nas primeiras 5 linhas.")


def parse_colaboradores_xlsx(file_bytes: bytes) -> list[dict]:
    """Lê a aba 'Colaboradores'. Suporta o formato simples (9 colunas) e o formato
    completo com blocos de férias."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    if COL_SHEET not in wb.sheetnames:
        raise ValueError(f"Aba '{COL_SHEET}' não encontrada no arquivo")
    ws = wb[COL_SHEET]

    header_row, col_map = _detect_colab_header(ws)
    data_start = header_row + 1
    has_ferias = ws.max_column >= 11

    def cell(row, campo):
        col = col_map.get(campo)
        return ws.cell(row=row, column=col).value if col else None

    registros = []
    for row in range(data_start, ws.max_row + 1):
        nome = _to_str(cell(row, "nome"))
        cpf = _to_str(cell(row, "cpf"))
        if not nome or not cpf:
            continue

        ferias = []
        if has_ferias:
            for inicio_col, termino_col, dias_col in FERIAS_BLOCKS:
                data_inicio = _to_date(ws.cell(row=row, column=inicio_col).value)
                data_termino = _to_date(ws.cell(row=row, column=termino_col).value)
                if data_inicio is None or data_termino is None:
                    continue
                dias_tirados_raw = ws.cell(row=row, column=dias_col).value
                dias_tirados = _to_int(dias_tirados_raw)
                if dias_tirados is None:
                    dias_tirados = max((data_termino - data_inicio).days, 0)
                ferias.append({
                    "ano": data_inicio.year,
                    "data_inicio": data_inicio,
                    "data_fim": data_termino,
                    "dias_tirados": dias_tirados,
                    "dias_direito": DIAS_DIREITO_DEFAULT,
                })

        registros.append({
            "_linha": row,
            "nome": nome,
            "cpf": cpf,
            "cargo": _to_str(cell(row, "cargo")),
            "salario": _to_float(cell(row, "salario")),
            "data_nascimento": _to_date(cell(row, "data_nascimento")),
            "endereco_completo": _to_str(cell(row, "endereco")),
            "data_admissao": _to_date(cell(row, "data_admissao")),
            "data_desligamento": _to_date(cell(row, "data_desligamento")),
            "ferias": ferias,
        })

    return registros


# Aliases para detecção dinâmica de colunas do fluxo de caixa
CONTAS_HEADER_ALIASES = [
    ("data",       ["data", "date", "dt"]),
    ("descricao",  ["descrição", "descricao", "historico", "histórico", "lançamento", "lancamento"]),
    ("debito",     ["débito", "debito", "saída", "saida", "debit", "saídas", "saidas"]),
    ("credito",    ["crédito", "credito", "entrada", "entradas", "credit"]),
]


def _detect_contas_header(ws) -> tuple[int, dict] | None:
    """Varre as primeiras 10 linhas acumulando colunas encontradas.

    Suporta cabeçalhos espalhados por múltiplas linhas via células mescladas
    (ex.: 'Data'/'Descrição' na linha 2 e 'Débito' na linha 4 do modelo padrão).
    Retorna (última_linha_de_cabeçalho, col_map).
    """
    col_map: dict[str, int] = {}
    last_header_row = 0

    for header_row in range(1, 11):
        row_vals = {
            c: (str(ws.cell(row=header_row, column=c).value or "")).strip().lower()
            for c in range(1, min(ws.max_column + 1, 20))
        }
        found_any = False
        for campo, aliases in CONTAS_HEADER_ALIASES:
            if campo not in col_map:
                for col, val in row_vals.items():
                    if val in aliases:
                        col_map[campo] = col
                        found_any = True
                        break
        if found_any:
            last_header_row = header_row

        if "debito" in col_map and ("data" in col_map or "descricao" in col_map):
            return last_header_row, col_map

    return None


# ==================== PARSING: Contas a Pagar ====================
def parse_contas_xlsx(file_bytes: bytes) -> list[dict]:
    """Lê todas as abas do arquivo de Fluxo de Caixa buscando lançamentos.

    Detecta automaticamente o cabeçalho e as colunas (data, descrição, débito).
    Importa APENAS linhas com valor na coluna débito > 0.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    registros = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        result = _detect_contas_header(ws)
        if result is None:
            continue
        header_row, col_map = result

        col_data      = col_map.get("data", CONTAS_COLS["data"])
        col_descricao = col_map.get("descricao", CONTAS_COLS["descricao"])
        col_debito    = col_map.get("debito", CONTAS_COLS["debito"])

        for row in range(header_row + 1, ws.max_row + 1):
            data_val  = _to_date(ws.cell(row=row, column=col_data).value)
            descricao = _to_str(ws.cell(row=row, column=col_descricao).value)

            # Linha sem descrição: pular
            if not descricao:
                continue

            # Importa apenas linhas COM valor de débito
            debito = _to_float(ws.cell(row=row, column=col_debito).value)
            if not debito or debito <= 0:
                continue

            # Sem data válida: pular (exclui linhas de totalização naturalmente)
            if data_val is None:
                continue

            from app.services.categorias_contas import inferir_de_descricao
            cat, sub = inferir_de_descricao(descricao)

            registros.append({
                "_aba": sheet_name,
                "_linha": row,
                "data_vencimento": None,
                "descricao": descricao,
                "categoria": cat,
                "subcategoria": sub,
                "valor": debito,
                "pago": True,
                "data_pagamento": data_val,
            })

    return registros


def _inferir_centro_custo(descricao: str) -> str:
    """Compat: retorna só categoria (legado). Preferir inferir_de_descricao."""
    from app.services.categorias_contas import inferir_de_descricao
    cat, _sub = inferir_de_descricao(descricao)
    return cat


# ==================== EXPORT: NFs ====================
def preencher_template_nfs(nfs: list[dict]) -> bytes:
    """Preenche a aba 'Entradas' do template com as NFs (em memória).

    Origem / Imposto / Data ent. pgto são acrescentados após Placement.
    Caixa não é exportada nesta planilha (019).
    """
    wb = openpyxl.load_workbook(TEMPLATE_NFS_PATH)
    ws = wb[NF_SHEET]

    col_origem = max(NF_COLS.values()) + 1
    col_imposto = col_origem + 1
    col_ent_pgto = col_imposto + 1
    ws.cell(row=NF_HEADER_ROW, column=col_origem, value="Origem")
    ws.cell(row=NF_HEADER_ROW, column=col_imposto, value="Imposto")
    ws.cell(row=NF_HEADER_ROW, column=col_ent_pgto, value="Data ent. pgto")

    row = NF_DATA_START_ROW
    for nf in nfs:
        ws.cell(row=row, column=NF_COLS["numero"],         value=nf.get("numero"))
        ws.cell(row=row, column=NF_COLS["razao_social"],   value=nf.get("razao_social"))
        ws.cell(row=row, column=NF_COLS["posicao"],        value=nf.get("posicao"))
        ws.cell(row=row, column=NF_COLS["valor_bruto"],    value=nf.get("valor_bruto"))
        ws.cell(row=row, column=NF_COLS["valor_liquido"],  value=nf.get("valor_liquido"))
        ws.cell(row=row, column=NF_COLS["data_emissao"],   value=nf.get("data_emissao"))
        ws.cell(row=row, column=NF_COLS["data_vencimento"], value=nf.get("data_vencimento"))
        ws.cell(row=row, column=NF_COLS["data_pagamento"], value=nf.get("data_pagamento"))
        ws.cell(row=row, column=NF_COLS["lead"],           value=nf.get("lead_nome"))
        ws.cell(row=row, column=NF_COLS["conducao"],       value=nf.get("conducao_nome"))
        ws.cell(row=row, column=NF_COLS["placement"],      value=nf.get("placement_nome"))
        orig = nf.get("origem") or "maggo"
        ws.cell(row=row, column=col_origem, value="Manual" if orig == "manual" else "Maggo")
        ws.cell(row=row, column=col_imposto, value=nf.get("valor_imposto"))
        ws.cell(row=row, column=col_ent_pgto, value=nf.get("data_ent_pgto"))
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ==================== EXPORT: Colaboradores ====================
def preencher_template_colaboradores(colaboradores: list[dict]) -> bytes:
    """Preenche a aba 'Colaboradores' do template simplificado (sem férias)."""
    wb = openpyxl.load_workbook(TEMPLATE_COLAB_PATH)
    ws = wb[COL_SHEET]

    row = COL_DATA_START_ROW
    for i, colab in enumerate(colaboradores, start=1):
        ws.cell(row=row, column=1,                         value=i)
        ws.cell(row=row, column=COLAB_COLS["nome"],        value=colab.get("nome"))
        ws.cell(row=row, column=COLAB_COLS["cpf"],         value=colab.get("cpf"))
        ws.cell(row=row, column=COLAB_COLS["cargo"],       value=colab.get("cargo"))
        ws.cell(row=row, column=COLAB_COLS["salario"],     value=colab.get("salario"))
        ws.cell(row=row, column=COLAB_COLS["data_nascimento"], value=colab.get("data_nascimento"))
        ws.cell(row=row, column=COLAB_COLS["endereco"],    value=colab.get("endereco_completo"))
        ws.cell(row=row, column=COLAB_COLS["data_admissao"], value=colab.get("data_admissao"))
        ws.cell(row=row, column=COLAB_COLS["data_desligamento"], value=colab.get("data_desligamento"))
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ==================== EXPORT: Contas a Pagar ====================
def preencher_template_contas(contas: list[dict], mes: int | None = None, ano: int | None = None) -> bytes:
    """Preenche a aba de fluxo de caixa do template com as contas (em memória).

    `contas` é lista de dicts com: data_vencimento, descricao, valor, pago, data_pagamento.
    Usa a única aba do modelo_contas.xlsx e atualiza o título com mês/ano fornecidos.
    """
    wb = openpyxl.load_workbook(TEMPLATE_CONTAS_PATH)
    ws = wb.active  # única aba do template

    # Atualiza título com mês/ano se fornecidos
    if mes and ano:
        mes_str = MESES_NOME.get(mes, str(mes))
        ws.cell(row=1, column=1, value=f"Fluxo de Caixa - Ocean - Conta de movimentação - {mes_str} {ano}")

    row = CONTAS_DATA_START_ROW
    for conta in contas:
        data = conta.get("data_pagamento") or conta.get("data_vencimento")
        ws.cell(row=row, column=CONTAS_COLS["data"],       value=data)
        ws.cell(row=row, column=CONTAS_COLS["descricao"],  value=conta.get("descricao"))
        ws.cell(row=row, column=CONTAS_COLS["forma_pgto"], value="Pix" if conta.get("pago") else "")
        ws.cell(row=row, column=CONTAS_COLS["debito"],     value=conta.get("valor"))
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
